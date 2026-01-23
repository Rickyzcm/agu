from datetime import datetime

import baostock as bs
from openpyxl import load_workbook
import threading
import time

## 股票信息的设计
class Stock:
    def __init__(self,code,name,sheet_name,start_date,end_date):
        self.code = code
        self.name = name
        self.sheet_name = sheet_name
        self.start_date = start_date
        self.end_date =end_date
        self.data_list=[]

    def set_sheet_name(self,sheet_name):
        self.sheet_name = sheet_name

    def set_data_list(self,data_list):
        self.data_list = data_list ## 某一个gu历史K线数据的集合

def query_to_map_list(
        code_map,
        ):
    lock = threading.Lock()
    for key, value in code_map.items():
        with lock:
            print(f"从'{value.start_date}'开始,到'{value.end_date}'为止")

            rs = bs.query_history_k_data_plus(key,
                                              "date,code,open,high,low,close,preclose,volume,amount,adjustflag,turn,tradestatus,pctChg,isST",
                                              value.start_date, value.end_date,
                                              frequency="d", adjustflag="2")
            print('query_history_k_data_plus respond error_code:' + rs.error_code)
            print('query_history_k_data_plus respond  error_msg:' + rs.error_msg)

            #### 打印结果集 ####
            data_list = []
            while (rs.error_code == '0') & rs.next():
                # 获取一条记录，将记录合并在一起
                data_list.append(rs.get_row_data())
            # result = pd.DataFrame(data_list, columns=rs.fields)

            #### 结果集输出到csv文件 ####
            # result.to_csv(
            #     "/Users/hyperchain/Documents/out/history_A_stock_k_data" + value.name + "_" + value.start_date + "~" + value.end_date + ".csv",
            #     index=False)

            ### 修改结果集输出到指定文件的指定工作表中
            # with pd.ExcelWriter('existing.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            #     result.to_excel(writer, sheet_name=sheet_name, index=False,
            #                     startrow=writer.sheets[sheet_name].max_row)
            # print(result)
            value.set_data_list(data_list)
## 将日k线数据追加到制定工作表中（）
def append_to_multiple_sheets_and_save_as_copy(original_file_path, new_file_path, code_map):
    """
        打开指定Excel文件，追加数据后另存为新文件。

        参数:
            original_file_path (str): 原始Excel文件路径。
            new_file_path (str): 要保存的新Excel文件路径。
            code_map (map): 要追加的数据（已经根据code归类好了），是一个列表的列表。每个子列表代表一行数据。
        """
    # 1. 加载原始工作簿
    wb = load_workbook(original_file_path)

    # 2. 遍历字典，向每个指定工作表追加数据
    for key, value in code_map.items():
        # 检查工作表是否存在
        if value.sheet_name in wb.sheetnames:
            ws = wb[value.sheet_name]
            print(f"正在向工作表 '{value.sheet_name}' 追加数据...")

            # 将数据行逐条追加到该工作表的末尾
            for row in value.data_list:
                ws.append(row)
            print(f"  已追加 {len(value.data_list)} 行数据。")
        else:
            print(f"警告: 工作簿中不存在名为 '{value.sheet_name}' 的工作表，已跳过。")

    # 3. 所有数据追加完成后，统一保存到新文件（生成副本）
    wb.save(new_file_path)
    print(f"\n所有数据已追加完成，并保存为新文件: {new_file_path}")


#### 登陆系统 ####
lg = bs.login()
# 显示登陆返回信息
print('login respond error_code:' + lg.error_code)
print('login respond  error_msg:' + lg.error_msg)

## 修改的源excel文件
original_file_path='/Users/hyperchain/Documents/个人持仓情况.xlsx'
new_file_path='/Users/hyperchain/Documents/个人持仓情况'+str(time.time())+'.xlsx'
## 获取当前最新日期
end_date = datetime.today().strftime( "%Y-%m-%d")
gu_list = [
    Stock('sh.603078','江化微','','2026-01-21',end_date),
    Stock('sz.000555','精工科技','','2026-01-21',end_date),
    Stock('sh.601127','赛力斯','','2026-01-21',end_date),
    Stock('sz.000977','浪潮信息','','2026-01-21',end_date),
    Stock('sz.002065','东华软件','','2025-01-01',end_date)
]
## 生成每个股票对应历史K线数据的工作表表名
for gu in gu_list:
    gu.set_sheet_name(gu.name+'历史K线数据') # 生成规则：拼接

#### 获取沪深A股历史K线数据 ####
# 详细指标参数，参见“历史行情指标参数”章节；“分钟线”参数与“日线”参数不同。“分钟线”不包含指数。
# 分钟线指标：date,time,code,open,high,low,close,volume,amount,adjustflag
# 周月线指标：date,code,open,high,low,close,volume,amount,adjustflag,turn,pctChg
codeMap = {gu.code:gu for gu in gu_list}

# 调用函数
query_to_map_list(codeMap)
append_to_multiple_sheets_and_save_as_copy(original_file_path,new_file_path,codeMap)
#### 登出系统 ####
bs.logout()
