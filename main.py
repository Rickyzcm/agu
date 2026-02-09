from datetime import datetime, timedelta

import baostock as bs
from openpyxl import load_workbook
import threading
import config


## 股票信息的数据对象设计
class Stock:

    def __init__(self, code: str, name: str, start_date: str, end_date: str):
        self.code = code
        self.name = name
        self.sheet_name = f'{self.name}历史K线数据'  ## 生成每个股票对应历史K线数据的工作表表名
        self.start_date = start_date
        self.end_date = end_date


    def set_sheet_name(self, sheet_name: str):
        self.sheet_name = sheet_name

    def set_rs(self, fields, data_list: []):
        self.fields = fields  # 列名
        self.data_list = data_list  ## 某一个gu历史K线数据的集合

## 读取excel文档并将最新数据日期+1成为查询标的的开始日期（一般放在查询数据前准备阶段执行）
def query_data_latest_date(
        code_map: dict,
        original_file_path: str,
        end_date: str):
    # 1. 加载原始工作簿
    wb = load_workbook(original_file_path)

    # 2. 遍历字典，向每个指定工作表追加数据：只会遍历查询目标股票的工作表，也就意味着不会访问非K线工作表
    for key, value in code_map.items():
        # 检查股票名称对应的工作表是否存在
        if value.sheet_name in wb.sheetnames:
            ws = wb[value.sheet_name]
            last_row = ws.max_row
            last_date = None
            try:
                # 若工作表存在且有数据
                if last_row > 1:
                    last_date = ws.cell(row=last_row, column=1).value
                    date_obj= None
                    # 取第一列数据内容，格式“”
                    if isinstance(last_date,datetime):
                        date_obj = last_date
                    elif isinstance(last_date,str):
                        date_obj = datetime.strptime(last_date, '%Y-%m-%d') # 解析成 datetime对象
                    next_day = date_obj + timedelta(days=1) # +1成第二天日期
                    value.start_date=next_day.strftime('%Y-%m-%d') # 该标的取值开始日期为最新数据的第二天
                    value.end_date=end_date # 该标的取值截止日期为输入的日期，一般就是执行程序时的系统日期
            except ValueError:
                print(f'{value.sheet_name}中行数为{last_row},第一个单元格的内容为{last_date}')
        else:
            value.start_date =config.__get_primary_data_start_date__()
            value.end_date = end_date

def query_to_map_list(
        code_map: dict,
):
    lock = threading.Lock()
    for key, value in code_map.items():
        with lock:
            print(f"查询'{value.name}'从'{value.start_date}'开始,到'{value.end_date}'为止")

            rs = bs.query_history_k_data_plus(key,
                                              "date,code,open,high,low,close,preclose,volume,amount,adjustflag,turn,tradestatus,pctChg,isST",
                                              value.start_date, value.end_date,
                                              frequency="d", adjustflag="2")
            print('query_history_k_data_plus respond error_code:' + rs.error_code)
            print('query_history_k_data_plus respond  error_msg:' + rs.error_msg)

            #### 打印结果集 ####
            data_list = []
            while (rs.error_code == '0') & rs.next():
                # 获取一条记录，将记录合并在一起
                data_list.append(rs.get_row_data())
            # result = pd.DataFrame(data_list, columns=rs.fields)
            value.set_rs(fields=rs.fields, data_list=data_list)

## 将日k线数据追加到制定工作表中（）
def append_to_multiple_sheets_and_save_as_copy(
        original_file_path: str,
        new_file_path: str,
        code_map):
    """
        打开指定Excel文件，追加数据后另存为新文件。

        参数:
            original_file_path (str): 原始Excel文件路径。
            new_file_path (str): 要保存的新Excel文件路径。
            code_map (map): 要追加的数据（已经根据code归类好了），是一个列表的列表。每个子列表代表一行数据。
        """
    # 1. 加载原始工作簿
    wb = load_workbook(original_file_path)

    # 2. 遍历字典，向每个指定工作表追加数据
    for key, value in code_map.items():
        # 检查工作表是否存在
        if value.sheet_name in wb.sheetnames:
            ws = wb[value.sheet_name]
            print(f"正在向工作表 '{value.sheet_name}' 追加数据...")
        else:
            print(f"警告: 工作簿中不存在名为 '{value.sheet_name}' 的工作表，将新建工作表。")
            ws = wb.create_sheet(title=value.sheet_name)
            # 第一行写列名
            ws.append(value.fields)
        # 将数据行逐条追加到该工作表的末尾
        for row in value.data_list:
            ws.append(row)
        print(f"  已追加 {len(value.data_list)} 行数据。")

    # 3. 所有数据追加完成后，统一保存到新文件（生成副本）
    wb.save(new_file_path)
    print(f"\n所有数据已追加完成，并保存为新文件: {new_file_path}")


# 主程序开始
## 读取本地配置
original_file_path = config.__get_original_file_path__()
new_file_path = config.__get_new_file_path__()

## 获取当前最新日期
current_date = datetime.today().strftime("%Y-%m-%d")
gu_list = [
    Stock('sz.000099', '中信海直', '2025-01-01', end_date=current_date),
    Stock('sz.000829', '天音控股', '2025-01-01', end_date=current_date),
    Stock('sz.001298', '好上好', '2025-01-01', end_date=current_date),
    Stock('sz.000636', '风华高科', '2025-01-01', end_date=current_date),
    Stock('sh.600879', '航天电子', '2026-01-17', end_date=current_date),
    Stock('sh.600487', '亨通光电', '2026-01-17', end_date=current_date),
    Stock('sh.603667', '五洲新春', '2026-01-17', end_date=current_date),

]

#### 获取沪深A股历史K线数据 ####
# 详细指标参数，参见“历史行情指标参数”章节；“分钟线”参数与“日线”参数不同。“分钟线”不包含指数。
# 分钟线指标：date,time,code,open,high,low,close,volume,amount,adjustflag
# 周月线指标：date,code,open,high,low,close,volume,amount,adjustflag,turn,pctChg
code_map = {gu.code: gu for gu in gu_list}
# 预处理一下每个标的查询时间区间（开始时间和结束时间）
query_data_latest_date(code_map=code_map,original_file_path = original_file_path,end_date=current_date)
#### 登陆系统 ####
lg = bs.login()
# 显示登陆返回信息
print('login respond error_code:' + lg.error_code)
print('login respond  error_msg:' + lg.error_msg)

# 调用函数
query_to_map_list(code_map)
#### 登出系统 ####
bs.logout()
## 数据处理进文件
append_to_multiple_sheets_and_save_as_copy(original_file_path, new_file_path, code_map)
